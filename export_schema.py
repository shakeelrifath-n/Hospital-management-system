import mysql.connector
from mysql.connector import Error
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Database connection settings
config = {
    'host': 'localhost',
    'user': 'root',
    'password': 'jee59',
    'database': 'hms'
}

def connect_to_database():
    """Connect to MySQL database"""
    try:
        connection = mysql.connector.connect(**config)
        if connection.is_connected():
            print("✅ Connected to MySQL database successfully")
            return connection
    except Error as e:
        print(f"❌ Error while connecting to MySQL: {e}")
        return None

def get_all_tables(connection):
    """Get list of all tables in database"""
    cursor = connection.cursor()
    cursor.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA = %s", 
                   (config['database'],))
    tables = [table[0] for table in cursor.fetchall()]
    cursor.close()
    return tables

def get_table_structure(connection, table_name):
    """Get column information for a table"""
    cursor = connection.cursor()
    cursor.execute(f"DESCRIBE `{table_name}`")
    columns = cursor.fetchall()
    cursor.close()
    return columns

def get_table_data(connection, table_name):
    """Get all data from a table"""
    cursor = connection.cursor(dictionary=True)
    cursor.execute(f"SELECT * FROM `{table_name}`")
    data = cursor.fetchall()
    cursor.close()
    return data

def create_schema_sheet(workbook, connection):
    """Create a sheet with database schema information"""
    ws = workbook.create_sheet("DATABASE SCHEMA", 0)
    
    # Title
    ws['A1'] = "Hospital Management System - Database Schema"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:F1')
    
    # Metadata
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Database: {config['database']}"
    
    # Get all tables
    tables = get_all_tables(connection)
    
    # Headers for schema table
    ws['A5'] = "Table Name"
    ws['B5'] = "Column Name"
    ws['C5'] = "Data Type"
    ws['D5'] = "Null"
    ws['E5'] = "Key"
    ws['F5'] = "Default"
    
    # Format headers
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}5'].fill = header_fill
        ws[f'{col}5'].font = header_font
        ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add schema information
    row = 6
    for table_name in sorted(tables):
        columns = get_table_structure(connection, table_name)
        for idx, column in enumerate(columns):
            col_name, col_type, col_null, col_key, col_default, col_extra = column
            
            if idx == 0:
                ws[f'A{row}'] = table_name
            ws[f'B{row}'] = col_name
            ws[f'C{row}'] = col_type
            ws[f'D{row}'] = col_null
            ws[f'E{row}'] = col_key
            ws[f'F{row}'] = str(col_default) if col_default is not None else ""
            
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    
    return workbook

def create_table_sheet(workbook, connection, table_name):
    """Create a sheet for a specific table with data"""
    try:
        # Get column structure
        columns = get_table_structure(connection, table_name)
        column_names = [col[0] for col in columns]
        
        # Get data
        data = get_table_data(connection, table_name)
        
        # Create worksheet
        ws = workbook.create_sheet(table_name)
        
        # Add title
        ws['A1'] = f"Table: {table_name}"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells(f'A1:{get_column_letter(len(column_names))}1')
        
        # Add column headers
        for col_idx, col_name in enumerate(column_names, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data rows
        for row_idx, row_data in enumerate(data, 3):
            for col_idx, col_name in enumerate(column_names, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(col_name, "")
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
        
        # Adjust column widths
        for col_idx, col_name in enumerate(column_names, 1):
            max_length = len(col_name)
            for row in ws.iter_rows(min_row=3, max_row=len(data)+2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Add summary info
        summary_row = len(data) + 4
        ws[f'A{summary_row}'] = "Total Records:"
        ws[f'B{summary_row}'] = len(data)
        ws[f'A{summary_row}'].font = Font(bold=True)
        
        print(f"✅ Created sheet for table: {table_name} ({len(data)} records)")
        
    except Exception as e:
        print(f"❌ Error creating sheet for {table_name}: {e}")

def main():
    """Main function to export database schema to Excel"""
    print("\n📊 Hospital Management System - Database Schema Export")
    print("=" * 60)
    
    # Connect to database
    connection = connect_to_database()
    if not connection:
        print("❌ Failed to connect to database")
        return
    
    try:
        # Create workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        # Create schema summary sheet
        print("\n📝 Creating schema summary sheet...")
        workbook = create_schema_sheet(workbook, connection)
        
        # Get all tables
        tables = get_all_tables(connection)
        print(f"\n📋 Found {len(tables)} tables")
        
        # Create a sheet for each table
        print("\n📊 Creating sheets for each table...")
        for table_name in sorted(tables):
            create_table_sheet(workbook, connection, table_name)
        
        # Save workbook
        output_file = "Hospital_Management_System_Database_Schema.xlsx"
        workbook.save(output_file)
        print(f"\n✅ Export completed successfully!")
        print(f"📁 File saved as: {output_file}")
        print(f"📍 Location: Current directory")
        
    except Exception as e:
        print(f"❌ Error during export: {e}")
    finally:
        if connection.is_connected():
            connection.close()
            print("\n✅ Database connection closed")

if __name__ == "__main__":
    main()
